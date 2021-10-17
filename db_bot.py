import openpyxl
import sqlite3

conn = sqlite3.connect('Sales_database.sqlite')
cur = conn.cursor()
cur.executescript('''
Create Table If NOT exists Products(
    Product_Key Integer PRIMARY KEY,
    Product_name Text Unique
);
Create Table if NOT exists Shops(
    Shop_id Text,
    Shop_name Text,
    Branch_code text unique,
    Branch_name text
    
);
create Table if not exists sales(
    Sales_order_no TEXT NOT NULL PRIMARY KEY UNIQUE,
    Date Text,
    Product_key integer,
    Shop_id Text,
    Branch_code Text,
    Unit_price Integer,
    Qty_sold Integer,
    Total_sales Integer 
)
''')
# The below script opens the product_codes workbook- The work book contains unique product keys and names
wb = openpyxl.load_workbook('product_codes.xlsx')
print(wb.sheetnames) # Checks the sheets within the workbook 
codes = wb['Product_codes']
codes_max = codes.max_row
codes_maxx = codes_max +1
# The below loop extracts the  product names and keys from the product_codes worksheet and inputs them into the database
for i in range (2,codes_maxx):
    product_name = (codes.cell(row = i, column = 1).value)
    product_Key = (codes.cell(row = i, column = 2).value)
    cur.execute('insert or ignore into Products (Product_Key,Product_name) values (?,?)', (product_Key, product_name))
    conn.commit()
# excel workbook of shop 1
wb = openpyxl.load_workbook('shop1.xlsx')
print('Fetching data from Shop 1 workbook')
sheet = wb['branch_1'] # The sheet contains sales data for shop 1 branch_1 
sheets = wb['shop_id'] # This sheet contains the shop_id for shop1
sheets_2 = wb['branches'] # This sheet contains the branches unique codes and  name that belong to shop1 
branch_code = (sheets_2.cell(row = 2, column = 1).value) # Assigns the branch code for shop 1 branch_1
branch = (sheets_2.cell(row = 2, column = 2).value)# Assigns the Branch name for branch_1 
shop_id = (sheets.cell(row = 2, column = 1).value)# Assigns the shop id uniqie to Shop 1
shop_name  = (sheets.cell (row = 2, column = 2).value) # Assigns the shop name  uniqie to Shop 1
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 1 branch_1 data into database')
# The below loop extracts the relevent information needed from the branch_1 worksheet for the database tables and inserts them to the tables
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
# Lines 30 to 65 is repeated below for shop1 branch_2, and 3 and for all the other shops and branches. 
print('Finished loading shop 1 branch_1 data into database')
sheet = wb['branch_2']  
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 3, column = 1).value) 
branch = (sheets_2.cell(row = 3, column = 2).value)
maxss = maxs+1
print('Inserting shop 1 branch_2 data into database')
# shop 1 branch_2
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 1 branch_2 data into database')
# Shop 1 branch_3
sheet = wb['branch_3'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 4, column = 1).value) 
branch = (sheets_2.cell(row = 4, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 1 branch_3 data into database')
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 1 branch_3 data into database')
# shop 2 branch_1
wb = openpyxl.load_workbook('shop2.xlsx')
print('Fetching data from Shop 2 workbook')
sheet = wb['branch_1'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 2, column = 1).value) 
branch = (sheets_2.cell(row = 2, column = 2).value)
shop_id = (sheets.cell(row = 2, column = 1).value)
shop_name  = (sheets.cell (row = 2, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 2 branch_1 data into database')
# shop 1
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 2 branch_1 data into database')
# shop 2 branch_2
sheet = wb['branch_2'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 3, column = 1).value) 
branch = (sheets_2.cell(row = 3, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 2 data branch_2 into database')
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 2 branch_2 data into database')
# shop 2 branch_3
sheet = wb['branch_3'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 4, column = 1).value) 
branch = (sheets_2.cell(row = 4, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 2 branch_3 data into database')
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 2 branch_3 data into database')
# shop 2 branch_4
sheet = wb['branch_4'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 5, column = 1).value) 
branch = (sheets_2.cell(row = 5, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 2 data branch_4 into database')
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 2 branch_4 data into database')
# shop 3 branch_1
wb = openpyxl.load_workbook('shop3.xlsx')
print('Fetching data from Shop 3 workbook')
sheet = wb['branch_1'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 2, column = 1).value) 
branch = (sheets_2.cell(row = 2, column = 2).value)
shop_id = (sheets.cell(row = 2, column = 1).value)
shop_name  = (sheets.cell (row = 2, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 3 Data branch_1 into database')
# shop 1
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 3 branch_1 data into database')
# shop 3 branch_2
sheet = wb['branch_2'] 
sheets = wb['shop_id'] 
sheets_2 = wb['branches'] 
branch_code = (sheets_2.cell(row = 3, column = 1).value) 
branch = (sheets_2.cell(row = 3, column = 2).value)
maxs = sheet.max_row
maxss = maxs+1
print('Inserting shop 3 Data branch_2 into database')
for i in range(2,maxss): 
    sales_order_no  = (sheet.cell(row = i,column=1).value)
    date = (sheet.cell(row = i, column = 2).value)
    product_Key = (sheet.cell(row=i, column=5).value)
    unit_price = (sheet.cell(row = i, column= 8).value)
    qty_sold = (sheet.cell(row = i, column = 7).value)
    total_sales = (sheet.cell(row = i, column= 9).value)
    cur.execute('insert or ignore into Shops (Shop_id,shop_Name, branch_code, branch_name) values (?,?,?,?)', (shop_id, shop_name,branch_code,branch))
    cur.execute('insert or replace into sales (Sales_order_no,Date,Product_key,Shop_id,Branch_code,Unit_price,Qty_sold,Total_sales) values (?,?,?,?,?,?,?,?)'
    ,(sales_order_no,date,product_Key,shop_id,branch_code,unit_price,qty_sold,total_sales))
    conn.commit()
print('Finished loading shop 3 branch_2 data into database')

print('Finished loading all data into databse')



